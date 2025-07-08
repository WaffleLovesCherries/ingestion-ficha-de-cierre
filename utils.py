import os
import io
import json
import re
import openpyxl as px

from typing import BinaryIO, Union

from multiprocessing import cpu_count
from multiprocessing.pool import ThreadPool

from operator import itemgetter
from dotenv import load_dotenv
from openpyxl import load_workbook
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File

from pandas import DataFrame
from typing import Callable, Optional, Any, Dict, List, Tuple
from openpyxl.workbook.workbook import Workbook
from typing import Literal
import pandas as pd

REPLACEMENTS = {
    'á': 'a', 
    'é': 'e', 
    'í': 'i', 
    'ó': 'o', 
    'ú': 'u', 
    'ü': 'u',
    'ñ': 'n'
}

def to_bool(value: str) -> bool:
    """Convert string to boolean.
    Accepts: 'true', 'false', '1', '0', 'yes', 'no', 'on', 'off' (case insensitive)
    Parameters:
        value (str): The string to convert to boolean.
    Returns:
        out (bool): The converted boolean value.
    Raises:
        ValueError: If the string cannot be converted to a boolean.
    """
    if isinstance(value, bool): return value
    if value.lower() in ('true', '1', 'yes', 'on', 't', 'y'): return True
    elif value.lower() in ('false', '0', 'no', 'off', 'f', 'n'): return False
    else: raise ValueError(f"Cannot convert '{value}' to boolean")

def clean_str(value: str, allow_chars: str|list = None) -> str:
    """Trim a string to remove unwanted characters.
    Parameters:
        value (str): The string to trim.
        allow_chars (str|list): Characters to keep in the string. If None, only alphanumeric characters are kept.
    Returns:
        out (str): The trimmed string.
    """
    if allow_chars is None:
        allow_chars = ''
    elif isinstance(allow_chars, str):
        allow_chars = list(allow_chars)
    
    pattern = f"[^a-záéíóúüñ0-9{''.join(allow_chars)}]"
    cleaned = re.sub(pattern, '', value.lower())

    for orig, repl in REPLACEMENTS.items():
        cleaned = cleaned.replace(orig, repl)
    return cleaned
    
def load_environment_variables():
    """Load environment variables from *.env* file, and optionally from a secret path if `BIG_BROTHER_WATCHING` is set.
    This is useful for loading sensitive information like API keys or database credentials by protecting them from GitHub Copilot.
    Raises:
        ValueError: If the string cannot be converted to a boolean.
    """
    load_dotenv(override=True)
    if os.getenv('BIG_BROTHER_WATCHING'):
        load_dotenv(os.getenv('SECRET_PATH'), override=True)


def setter(data: DataFrame, idx: Any) -> Callable[[str, Any], None]:
    """
    Returns a setter function to update a specific row in a DataFrame.

    Parameters:
        data (DataFrame): The DataFrame to update.
        idx (Any): The index of the row to update.

    Returns:
        Callable[[str, Any], None]: A function that sets a value in the specified row and column.
    """
    def _set(column: str, value: Any) -> None:
        data.loc[idx, column] = value
    return _set

def valid_code(code: Any) -> bool:
    """
    Checks if a code string is valid based on predefined prefixes and length.

    Parameters:
        code (Any): The code to validate.

    Returns:
        bool: True if the code is valid, False otherwise.
    """
    valid_prefixes = ['EDU', 'INC', 'GEO', 'MA', 'SAL', 'DT', 'AFI']
    code_str = str(code).strip()
    return any(code_str.startswith(prefix) for prefix in valid_prefixes) and len(code_str) > 4

def wbs_check(workbook: Workbook, set_data: Callable[[str, Any], None]) -> Optional[Dict[str, Any]]:
    """
    Extracts relevant information from a WBS Excel workbook.

    Parameters:
        workbook (Workbook): The openpyxl Workbook object.
        set_data (Callable[[str, Any], None]): Setter function to update DataFrame.

    Returns:
        Optional[Dict[str, Any]]: Dictionary with extracted fields or None if not found.
    """
    clean_sheetnames = {clean_str(sheet): sheet for sheet in workbook.sheetnames}

    # TODO FULL FIND FOR CODE
    if 'portada' not in clean_sheetnames: return None
    codigo = workbook[clean_sheetnames['portada']]['H27'].value
    if not valid_code(codigo):
        if 'conexioncronograma' not in clean_sheetnames:
            print(f"WARNING!!! code: {codigo} FAILED SECOND CHECK")
            print(clean_sheetnames)
            return None
        codigo = workbook[clean_sheetnames['conexioncronograma']]['A2'].value
        if not valid_code:
            print(f"WARNING!!! code: {codigo} FAILED SECOND CHECK")
            print(workbook[clean_sheetnames['portada']].values)
            return None

    if codigo is None: return None
    set_data('Codigo', codigo)
    if 'fichacierre' not in clean_sheetnames: return None
    ficha_sheet = workbook[clean_sheetnames['fichacierre']]
    rows: List[List[str]] = []
    for row in ficha_sheet.values:
        if any(cell is not None for cell in row):
            row_values = [str(cell) for cell in row if cell is not None and str(cell).strip() != '']
            if row_values: rows.append(row_values)
    if not rows: return None
    ficha: Dict[str, Any] = {'Codigo':codigo}
    for row in rows:
        if len(row) < 2: continue
        header = clean_str(row[0])
        value = ''.join(row[1:])
        match header:
            case 'retos': ficha['Retos'] = value
            case 'accionesdemitigacion': ficha['AccionesDeMitigacion'] = value
            case 'leccionesaprendidas': ficha['LeccionesAprendidas'] = value
    if len(ficha) <= 1: return None
    set_data('FichaCierre', True)
    return ficha

class SharepointLoader (ClientContext):
    """
    SharepointLoader is a class that extends ClientContext to interact with SharePoint.
    It initializes the client context with the SharePoint URL and authentication details from environment variables.
    """
    def __init__(self, environment=None, allow_ntlm=False, browser_mode=False):
        """
        Initialize the SharepointLoader with the SharePoint URL and authentication details.
        Parameters:
            environment (str): The environment to use, defaults to None.
            allow_ntlm (bool): Whether to allow NTLM authentication, defaults to False.
            browser_mode (bool): Whether to use browser mode for authentication, defaults to False.
        Raises:
            EnvironmentError: If required environment variables are missing or have invalid formats.
            ValueError: If the AUTH JSON cannot be parsed.
        """

        # Load environment variables
        load_environment_variables()
        base_url = os.getenv('SHAREPOINT_URL')
        auth_json = os.getenv('AUTH')

        # Validate required environment variables
        if not base_url or not auth_json:
            raise EnvironmentError("Missing required environment variables: SHAREPOINT_URL or AUTH")
        try:
            auth = json.loads(auth_json)
        except json.JSONDecodeError as e:
            raise ValueError(f"Failed to parse AUTH JSON: {e}")
        
        # Validate authentication parameters
        auth_context = AuthenticationContext(base_url)
        user = auth.get('username', 'unknown')[:4] + '***'  # Mask username for security
        print(f"Acquiring token for user: {user}")
        if not auth_context.acquire_token_for_user(**auth):
            raise PermissionError(f"Failed to acquire token for user {user}.")
        
        super().__init__(base_url, auth_context, environment, allow_ntlm, browser_mode)
        
    def get_files(self, folder_path: str, targets: list = []) -> list:
        """
        Get a list of files in a SharePoint folder.
        Parameters:
            folder_path (str): Path to the folder in SharePoint.
        Returns:
            out (list): List of file names in the specified folder.
        Raises:
            ValueError: If the folder path is empty.
            IOError: If the folder cannot be accessed.
        """
        if not folder_path:
            raise ValueError("Folder path cannot be empty.")
        
        try:
            if not targets:
                # Obtain folders and files by server relative URL
                files = self.web.get_folder_by_server_relative_url(folder_path).files
                folders = self.web.get_folder_by_server_relative_url(folder_path).folders
                self.load(files)
                self.load(folders)
                self.execute_query()

                # Get relevant file properties
                properties = ['UniqueId', 'Name', 'ServerRelativeUrl', 'TimeLastModified']
                filter = itemgetter(*properties)
                files = [dict(zip(properties, filter(file.properties))) for file in files if file.properties['Name'].lower().endswith(('.xlsm', '.xlsx', '.xls', '.csv'))]

                # Recursively load files in subfolders
                for folder in folders:
                    subfolder_files = self.get_files(folder.properties['ServerRelativeUrl'])
                    files.extend(subfolder_files)
                return files
            
            files = []
            # If targets are specified, load files from those subfolders
            for target in targets:
                subfolder_path = folder_path.rstrip('/') + '/' + target.lstrip('/')
                print(f"Loading files from subfolder: {target}")
                subfolder_files = self.get_files(subfolder_path)
                files.extend(subfolder_files)
            return files

        except Exception as e:
            raise IOError(f"Failed to access folder in SharePoint: {e}")

    def load_file(
        self,
        file_path: str,
        as_format: Literal['binary', 'workbook', 'dataframe'] = 'binary'
    ) -> BinaryIO | px.Workbook | DataFrame:
        """
        Load a file from the SharePoint site in the specified format.

        Parameters:
            file_path (str): Path to the file in SharePoint.
            as_format (Literal): Output format: 'binary', 'workbook', or 'dataframe'.
            **kwargs: Additional arguments for pandas.read_excel if as_format='dataframe'.

        Returns:
            BinaryIO | px.Workbook | DataFrame: The file content in the requested format.

        Raises:
            ValueError: If the file path is empty or as_format is invalid.
            IOError: If the file cannot be loaded from SharePoint.
        """
        if not file_path:
            raise ValueError("File path cannot be empty.")
        try:
            response = File.open_binary(self, file_path)
            buffer = io.BytesIO(response.content)
            if as_format == 'binary':
                return buffer
            elif as_format == 'workbook':
                return px.load_workbook(buffer, read_only=True, data_only=True)
            elif as_format == 'dataframe':
                data = px.load_workbook(buffer, read_only=True, data_only=True).active.values
                columns = next(data)
                data = pd.DataFrame(data, columns=columns)
                return data
            else:
                raise ValueError(f"Invalid as_format: {as_format}")
        except Exception as e:
            raise IOError(f"Failed to load file from SharePoint: {e}")
        """
        Load a file from the SharePoint site selected.
        Parameters:
            file_path (str): Path to the file in SharePoint.
            as_workbook (bool): If True, return an openpyxl Workbook; otherwise, return a BytesIO buffer.
        Returns:
            BinaryIO | px.Workbook: A BytesIO buffer containing the file content, or an openpyxl Workbook if as_workbook is True.
        Raises:
            ValueError: If the file path is empty.
            IOError: If the file cannot be loaded from SharePoint.
        """
        if not file_path:
            raise ValueError("File path cannot be empty.")
        
        try:
            response = File.open_binary(self, file_path)
            buffer = io.BytesIO(response.content)
            if as_workbook: return px.load_workbook(buffer, read_only=True, data_only=True)
            return buffer
        except Exception as e:
            raise IOError(f"Failed to load file from SharePoint: {e}")
    
    def save_file(self, file_path: str, buffer: Union[bytes, io.BytesIO]):
        """
        Save a file to the SharePoint site selected. If the file does not exist, it will be created.
        Parameters:
            file_path (str): Path to the file in SharePoint.
            buffer (bytes or BytesIO): The file content.
        Returns:
            out (bool): True if the file was saved successfully, False otherwise.
        Raises:
            ValueError: If the file path is empty.
            IOError: If the file cannot be saved to SharePoint.
        """
        if not file_path:
            raise ValueError("File path cannot be empty.")
        # Accept both bytes and BytesIO
        buffer = io.BytesIO(buffer) if isinstance(buffer, bytes) else buffer
        try:
            # Split the file_path into folder and filename
            folder_url, filename = file_path.rsplit('/', 1)
            folder = self.web.get_folder_by_server_relative_url(folder_url)
            # Upload file (will create or overwrite)
            folder.upload_file(filename, buffer.getvalue()).execute_query()
            return True
        except Exception as e:
            raise IOError(f"Failed to save file to SharePoint: {e}")

    def process_wbs(self, data: DataFrame) -> Tuple[DataFrame, DataFrame]:
        """
        Processes a DataFrame of WBS files, extracting information from each workbook.

        Parameters:
            data (DataFrame): DataFrame with at least a 'ServerRelativeUrl' column.

        Returns:
            Tuple[DataFrame, DataFrame]: 
                - Updated DataFrame with possible new columns.
                - DataFrame with extracted results from each workbook.
        """
        data_copy = data.copy()
        urls = list(data_copy['ServerRelativeUrl'])
        with ThreadPool(processes=cpu_count()) as pool:
            workbooks = pool.map(lambda url: self.load_file(url, as_format='workbook'), urls)
        results: List[Optional[Dict[str, Any]]] = []
        for idx, (workbook, row) in enumerate(zip(workbooks, data_copy.itertuples())):
            result = wbs_check(workbook, setter(data_copy, row.Index))
            results.append(result)
        filtered_results = [result for result in results if result is not None]
        results_df = DataFrame(filtered_results) if filtered_results else DataFrame()
        return data_copy, results_df
    
if __name__ == "__main__":
    loader = SharepointLoader()
    files = loader.get_files('/sites/MicrositioProyectosFSD/Documentos compartidos/', ['2. Gestión de Proyecto', '3. Proyectos cerrados'])
    print(len(files))