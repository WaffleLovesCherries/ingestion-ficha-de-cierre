from openpyxl.workbook.workbook import Workbook
from pandas import DataFrame
from typing import Any, Callable, Dict, List, Optional

from .text_utils import clean_str, valid_code

def setter(data: DataFrame, idx: Any) -> Callable[[str, Any], None]:
    """
    Returns a setter function to update a specific row in a DataFrame.

    Parameters:
        data (DataFrame): The DataFrame to update.
        idx (Any): The index of the row to update.

    Returns:
        Callable[[str, Any], None]: A function that sets a value in the specified row and column.
    """
    def _set(column: str, value: Any) -> None:
        data.loc[idx, column] = value
    return _set

def wbs_check(workbook: Workbook, set_data: Callable[[str, Any], None]) -> Optional[Dict[str, Any]]:
    """
    Extracts relevant information from a WBS Excel workbook.

    Parameters:
        workbook (Workbook): The openpyxl Workbook object.
        set_data (Callable[[str, Any], None]): Setter function to update DataFrame.

    Returns:
        Optional[Dict[str, Any]]: Dictionary with extracted fields or None if not found.
    """
    clean_sheetnames = {clean_str(sheet): sheet for sheet in workbook.sheetnames}

    # TODO FULL FIND FOR CODE
    if 'portada' not in clean_sheetnames: return None
    codigo = workbook[clean_sheetnames['portada']]['H27'].value
    if not valid_code(codigo):
        if 'conexioncronograma' not in clean_sheetnames:
            print(f"WARNING!!! code: {codigo} FAILED SECOND CHECK")
            print(clean_sheetnames)
            return None
        codigo = workbook[clean_sheetnames['conexioncronograma']]['A2'].value
        if not valid_code:
            print(f"WARNING!!! code: {codigo} FAILED SECOND CHECK")
            print(workbook[clean_sheetnames['portada']].values)
            return None

    if codigo is None: return None
    codigo = str(codigo).strip()
    set_data('Codigo', codigo)
    if 'fichacierre' not in clean_sheetnames: return None
    ficha_sheet = workbook[clean_sheetnames['fichacierre']]
    rows: List[List[str]] = []
    for row in ficha_sheet.values:
        if any(cell is not None for cell in row):
            row_values = [str(cell) for cell in row if cell is not None and str(cell).strip() != '']
            if row_values: rows.append(row_values)
    if not rows: return None
    ficha: Dict[str, Any] = {'Codigo':codigo}
    for row in rows:
        if len(row) < 2: continue
        header = clean_str(row[0])
        value = ''.join(row[1:])
        match header:
            case 'retos': ficha['Retos'] = value
            case 'accionesdemitigacion': ficha['AccionesDeMitigacion'] = value
            case 'leccionesaprendidas': ficha['LeccionesAprendidas'] = value
    if len(ficha) <= 1: return None
    set_data('FichaCierre', True)
    return ficha
